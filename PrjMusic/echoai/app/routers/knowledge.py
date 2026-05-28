"""
知识库管理路由。
提供文档上传、向量检索、文档删除等接口，仅管理员可用。
"""
import io
from typing import Optional

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile

from app.dependencies import CurrentUser, require_admin
from app.models.schemas import KbSearchRequest, KbUploadResponse
from app.services.vector_store import VectorStoreError, vector_store

router = APIRouter(prefix="/knowledge", tags=["knowledge-base"])


# 支持的文件类型与扩展名映射
SUPPORTED_MIME_TYPES = {
    "text/plain": "txt",
    "application/pdf": "pdf",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": "xlsx",
}

MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB


async def _extract_text(file: UploadFile) -> tuple[str, str]:
    """
    从上传文件中提取纯文本内容。

    :param file: FastAPI UploadFile
    :return: (文件内容, 原始文件名)
    :raises HTTPException: 文件类型不支持或解析失败
    """
    content_type = file.content_type or ""
    filename = file.filename or "unknown"

    # 优先使用声明的 content_type，fallback 到文件名后缀
    ext = SUPPORTED_MIME_TYPES.get(content_type)
    if not ext:
        if filename.lower().endswith(".txt"):
            ext = "txt"
        elif filename.lower().endswith(".pdf"):
            ext = "pdf"
        elif filename.lower().endswith(".xlsx"):
            ext = "xlsx"
        else:
            raise HTTPException(
                status_code=400,
                detail=f"不支持的文件类型: {content_type or filename}。仅支持 txt/pdf/xlsx",
            )

    # 读取文件内容
    raw_bytes = await file.read()
    if len(raw_bytes) > MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail="文件大小超过 10MB 限制")

    if ext == "txt":
        # 尝试 utf-8，失败则用 gbk
        for encoding in ("utf-8", "gbk", "latin-1"):
            try:
                return raw_bytes.decode(encoding), filename
            except UnicodeDecodeError:
                continue
        raise HTTPException(status_code=400, detail="文本文件编码无法识别")

    elif ext == "pdf":
        try:
            from pypdf import PdfReader
        except ImportError as e:
            raise HTTPException(status_code=500, detail="PDF 解析库未安装") from e

        try:
            reader = PdfReader(io.BytesIO(raw_bytes))
            texts = []
            for page in reader.pages:
                page_text = page.extract_text()
                if page_text:
                    texts.append(page_text)
            return "\n".join(texts), filename
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"PDF 解析失败: {e}") from e

    elif ext == "xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as e:
            raise HTTPException(status_code=500, detail="Excel 解析库未安装") from e

        try:
            wb = load_workbook(io.BytesIO(raw_bytes), data_only=True)
            texts = []
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                sheet_texts = []
                for row in sheet.iter_rows(values_only=True):
                    row_text = " ".join(str(cell) for cell in row if cell is not None)
                    if row_text.strip():
                        sheet_texts.append(row_text)
                if sheet_texts:
                    texts.append(f"--- Sheet: {sheet_name} ---\n" + "\n".join(sheet_texts))
            return "\n\n".join(texts), filename
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Excel 解析失败: {e}") from e

    raise HTTPException(status_code=400, detail="未知文件类型")


@router.get("", response_model=dict)
async def kb_list(
    page_num: int = 1,
    page_size: int = 20,
    user: CurrentUser = Depends(require_admin),
):
    """
    分页查询知识库文档列表。
    仅管理员可用。
    """
    docs, total = await vector_store.list_documents(page_num, page_size)
    return {
        "code": 200,
        "msg": "ok",
        "data": {
            "total": total,
            "records": docs,
        },
    }


@router.post("", response_model=dict)
async def kb_upload(
    file: UploadFile = File(..., description="上传文件（txt/pdf/xlsx）"),
    title: Optional[str] = Form(default=None, description="文档标题，为空则使用文件名"),
    user: CurrentUser = Depends(require_admin),
):
    """
    上传文档到知识库。
    仅管理员可用（role >= 2）。
    文档将被切分、生成 Embedding 后存入 echovector。
    """
    # 提取文本
    content, filename = await _extract_text(file)

    if not content.strip():
        raise HTTPException(status_code=400, detail="文件内容为空")

    doc_title = title.strip() if title and title.strip() else filename

    # 写入向量库
    try:
        result = await vector_store.add_document(
            title=doc_title,
            content=content,
            source_type="manual_upload",
            metadata={
                "original_filename": filename,
                "uploaded_by": user.user_id,
                "uploaded_by_name": user.username,
            },
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"文档入库失败: {e}") from e

    return {
        "code": 200,
        "msg": "上传成功",
        "data": {
            "doc_id": result["doc_id"],
            "title": doc_title,
            "total_chunks": result["total_chunks"],
        },
    }


@router.post("/search", response_model=dict)
async def kb_search(
    request: KbSearchRequest,
    user: CurrentUser = Depends(require_admin),
):
    """
    知识库向量检索（测试接口）。
    仅管理员可用。
    """
    try:
        results = await vector_store.similarity_search(
            query=request.query,
            top_k=request.top_k,
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"检索失败: {e}") from e

    return {
        "code": 200,
        "msg": "ok",
        "data": {
            "query": request.query,
            "total": len(results),
            "results": results,
        },
    }


@router.delete("/{doc_id}", response_model=dict)
async def kb_delete_document(
    doc_id: str,
    user: CurrentUser = Depends(require_admin),
):
    """
    软删除知识库中的指定文档。
    仅管理员可用。
    """
    deleted_count = await vector_store.delete_by_doc_id(doc_id)
    if deleted_count == 0:
        raise HTTPException(status_code=404, detail="文档不存在或已被删除")

    return {
        "code": 200,
        "msg": "删除成功",
        "data": {"doc_id": doc_id, "deleted_chunks": deleted_count},
    }


@router.get("/{doc_id}", response_model=dict)
async def kb_get_document(
    doc_id: str,
    user: CurrentUser = Depends(require_admin),
):
    """
    查看指定文档的分块详情。
    仅管理员可用。
    """
    chunks = await vector_store.get_document_chunks(doc_id)
    if not chunks:
        raise HTTPException(status_code=404, detail="文档不存在或已被删除")

    return {
        "code": 200,
        "msg": "ok",
        "data": {
            "doc_id": doc_id,
            "title": chunks[0]["title"] if chunks else None,
            "total_chunks": chunks[0]["total_chunks"] if chunks else 0,
            "chunks": chunks,
        },
    }
